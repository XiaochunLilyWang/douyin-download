#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
批量下载抖音链接的图片或视频。
数据来源：本地 Excel 文件，筛选来源渠道为「抖音 APP」的行。

处理流程（每条链接）：
  1. 尝试图集下载 → 成功则保存到 <output_dir>/<video_id>/images/
  2. 图集失败 → 尝试视频下载 → 成功则保存到 <output_dir>/<video_id>/video.mp4
  3. 两者都失败 → 标记为「无效」

子文件夹命名：使用抖音视频 ID（从 URL 中提取，与 Excel 中未来填写的 case ID 对应）

用法：
    python batch-download-douyin.py <input.xlsx> [output_dir]

依赖：
    pip install openpyxl requests
    Node.js 18+（用于调用 ../scripts/ 下的 JS 脚本）
"""

import json
import os
import re
import subprocess
import sys
import time
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# ——— 配置 ———
SOURCE_CHANNEL_VALUE = "抖音 APP"   # Excel 来源渠道列的值（中间有空格）
REQUEST_DELAY = 1.5                  # 每两次下载之间的间隔（秒）
NODE_TIMEOUT = 30                    # 每条 node 命令的超时时间（秒）

# JS 脚本目录（与本文件同目录）
SCRIPTS_DIR = Path(__file__).parent


def extract_video_id(url: str) -> str:
    """从抖音 URL 中提取视频 ID，作为子文件夹名称。"""
    for pattern in [r"video/([^/?]+)", r"note/([^/?]+)", r"modal_id=([0-9]+)"]:
        m = re.search(pattern, url)
        if m:
            return m.group(1)
    # 兜底：对 URL 做简单哈希，保证文件夹名合法
    return re.sub(r"[^\w-]", "_", url)[-40:]


def run_node(script_name: str, *args, timeout: int = NODE_TIMEOUT):
    """
    在 scripts/ 目录下运行 node 脚本。
    返回 (returncode, stdout, stderr)
    """
    cmd = [
        "node",
        str(SCRIPTS_DIR / script_name),
        *[str(a) for a in args],
    ]
    try:
        result = subprocess.run(
            cmd,
            cwd=str(SCRIPTS_DIR),
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=timeout,
        )
        return result.returncode, result.stdout, result.stderr
    except subprocess.TimeoutExpired:
        return -1, "", f"超时（>{timeout}s）"
    except FileNotFoundError:
        return -1, "", "找不到 node 命令，请确认 Node.js 已安装并在 PATH 中"


def try_image_album(url: str, images_dir: Path):
    """
    尝试图集下载。
    返回 (success, image_count, err_msg)
    """
    images_dir.mkdir(parents=True, exist_ok=True)
    code, stdout, stderr = run_node("extract-douyin-images.js", url, str(images_dir))

    if code == 0:
        # 从 stdout JSON 获取下载数量
        count = 0
        for line in stdout.splitlines():
            line = line.strip()
            if line.startswith("{"):
                try:
                    data = json.loads(line)
                    count = data.get("downloaded_images", 0)
                    break
                except Exception:
                    pass
        # 兜底：数目录文件
        if count == 0:
            count = len(list(images_dir.glob("image-*.*")))
        if count > 0:
            return True, count, ""
        else:
            return False, 0, "图集下载完成但未找到图片文件"
    else:
        err = (stderr.strip().split("\n")[-1] if stderr.strip() else "未知错误")
        return False, 0, err


def try_video(url: str, video_path: Path):
    """
    尝试视频下载。
    返回 (success, err_msg)
    """
    video_path.parent.mkdir(parents=True, exist_ok=True)
    code, stdout, stderr = run_node("parse-douyin-video.js", url, str(video_path))

    if code == 0 and video_path.exists() and video_path.stat().st_size > 0:
        return True, ""
    else:
        err = (stderr.strip().split("\n")[-1] if stderr.strip() else "未知错误")
        if "无法获取有效的视频播放地址" in stderr or "未找到包含视频数据" in stderr:
            err = "未找到视频播放地址"
        elif code == 0 and (not video_path.exists() or video_path.stat().st_size == 0):
            err = "视频文件未生成或为空"
        return False, err


def writeback_excel(input_path: str, results: list) -> None:
    """
    将下载结果回写到原 Excel：
    - 「链接是否有效」：是 / 否
    - 「媒体类型」：图集 / 视频 / （无效则留空）
    如「媒体类型」列不存在，自动在末尾新增。
    """
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    url_col = headers.index("原文URL") + 1

    # 「链接是否有效」必须已存在
    if "链接是否有效" not in headers:
        print("  ⚠ 原 Excel 中未找到「链接是否有效」列，跳过回写。")
        return
    valid_col = headers.index("链接是否有效") + 1

    # 「媒体类型」不存在则新增
    if "媒体类型" in headers:
        media_col = headers.index("媒体类型") + 1
    else:
        media_col = len(headers) + 1
        ws.cell(row=1, column=media_col, value="媒体类型").font = Font(bold=True)

    # 建立 URL → 结果 映射
    url_map = {r["url"]: r for r in results if r.get("url")}

    updated = 0
    for row in ws.iter_rows(min_row=2):
        url = str(row[url_col - 1].value or "").strip()
        if url not in url_map:
            continue
        r = url_map[url]
        row[valid_col - 1].value = "否" if r["type"] == "无效" else "是"
        ws.cell(row=row[0].row, column=media_col).value = (
            "" if r["type"] == "无效" else r["type"]
        )
        updated += 1

    try:
        wb.save(input_path)
        print(f"\n原 Excel 已同步：更新 {updated} 行（链接是否有效 + 媒体类型）→ {input_path}")
    except PermissionError:
        print(f"\n⚠ 无法写入原 Excel（文件可能正在被打开）：{input_path}")
        print("  请关闭该文件后手动重跑，或先查看 download_results.xlsx。")


def run(input_path: str, output_dir: str | None = None) -> None:
    input_path = Path(input_path)

    if output_dir is None:
        output_dir = input_path.parent / f"douyin_downloads_{input_path.stem}"
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # ——— 读取 Excel ———
    wb = openpyxl.load_workbook(str(input_path))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col_idx = {name: i for i, name in enumerate(headers)}

    url_col = col_idx.get("原文URL")
    channel_col = col_idx.get("来源渠道")
    title_col = col_idx.get("标题")
    case_id_col = col_idx.get("case ID")

    if url_col is None or channel_col is None:
        print("ERROR: 找不到「原文URL」或「来源渠道」列，请检查表头。")
        sys.exit(1)

    targets = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[channel_col] == SOURCE_CHANNEL_VALUE:
            targets.append({
                "case_id": row[case_id_col] if case_id_col is not None else None,
                "url": str(row[url_col] or "").strip(),
                "title": str(row[title_col] or "").strip() if title_col is not None else "",
            })

    total = len(targets)
    print(f"找到 {total} 条「{SOURCE_CHANNEL_VALUE}」链接")
    print(f"输出目录：{output_dir}\n")
    print("=" * 60)

    results = []

    for i, item in enumerate(targets, 1):
        url = item["url"]
        label = f"[{i:2d}/{total}]"

        if not url:
            print(f"\n{label} 跳过（URL 为空）")
            results.append({**item, "idx": i, "video_id": "", "type": "无效", "files": "", "note": "URL为空"})
            continue

        video_id = extract_video_id(url)
        item_dir = output_dir / video_id

        print(f"\n{label} {video_id}")
        print(f"        {url}")

        # ——— 步骤 1：尝试图集 ———
        images_dir = item_dir / "images"
        ok, count, err = try_image_album(url, images_dir)

        if ok:
            print(f"  → 图集 ✓  {count} 张图片")
            results.append({
                **item, "idx": i, "video_id": video_id,
                "type": "图集",
                "files": str(images_dir),
                "note": f"下载 {count} 张图片",
            })
        else:
            print(f"  → 图集失败：{err}")
            # 清理空目录
            for d in [images_dir, item_dir]:
                if d.exists() and not list(d.iterdir()):
                    d.rmdir()

            # ——— 步骤 2：尝试视频 ———
            video_path = item_dir / "video.mp4"
            v_ok, v_err = try_video(url, video_path)

            if v_ok:
                size_mb = video_path.stat().st_size / 1024 / 1024
                print(f"  → 视频 ✓  {size_mb:.1f} MB")
                results.append({
                    **item, "idx": i, "video_id": video_id,
                    "type": "视频",
                    "files": str(video_path),
                    "note": f"{size_mb:.1f} MB",
                })
            else:
                print(f"  → 视频失败：{v_err}  →【无效】")
                # 清理
                if video_path.exists():
                    video_path.unlink()
                for d in [item_dir]:
                    if d.exists() and not list(d.iterdir()):
                        d.rmdir()

                results.append({
                    **item, "idx": i, "video_id": video_id,
                    "type": "无效",
                    "files": "",
                    "note": f"图集: {err} | 视频: {v_err}",
                })

        if i < total:
            time.sleep(REQUEST_DELAY)

    # ——— 回写原 Excel：链接是否有效 + 媒体类型 ———
    writeback_excel(str(input_path), results)

    # ——— 写出结果 Excel ———
    out_xlsx = output_dir / "download_results.xlsx"
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "下载结果"

    headers_out = ["序号", "case_id", "视频ID（文件夹名）", "原文URL", "标题", "类型", "文件路径", "备注"]
    out_ws.append(headers_out)
    for cell in out_ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    FILL = {
        "图集": PatternFill("solid", fgColor="D6EAD6"),
        "视频": PatternFill("solid", fgColor="D6E4F0"),
        "无效": PatternFill("solid", fgColor="FAD7D7"),
    }

    for r in results:
        out_ws.append([
            r["idx"],
            r.get("case_id") or "",
            r.get("video_id", ""),
            r.get("url", ""),
            r.get("title", ""),
            r["type"],
            r["files"],
            r["note"],
        ])
        fill = FILL.get(r["type"])
        if fill:
            for cell in out_ws[out_ws.max_row]:
                cell.fill = fill

    out_ws.column_dimensions["A"].width = 6
    out_ws.column_dimensions["B"].width = 14
    out_ws.column_dimensions["C"].width = 22
    out_ws.column_dimensions["D"].width = 55
    out_ws.column_dimensions["E"].width = 38
    out_ws.column_dimensions["F"].width = 8
    out_ws.column_dimensions["G"].width = 55
    out_ws.column_dimensions["H"].width = 40

    out_wb.save(str(out_xlsx))

    # ——— 汇总 ———
    albums = sum(1 for r in results if r["type"] == "图集")
    videos = sum(1 for r in results if r["type"] == "视频")
    invalid = sum(1 for r in results if r["type"] == "无效")

    print("\n" + "=" * 60)
    print(f"完成！共 {total} 条")
    print(f"  图集: {albums} 条")
    print(f"  视频: {videos} 条")
    print(f"  无效: {invalid} 条")
    print(f"  结果表格: {out_xlsx}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python batch-download-douyin.py <input.xlsx> [output_dir]")
        sys.exit(1)
    run(sys.argv[1], sys.argv[2] if len(sys.argv) >= 3 else None)
